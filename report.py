import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
import config

# червоний напівжирний шрифт для поганих оцінок/коментарів
RED = Font(color="CC0000", bold=True)

COLS = {
    "date": 1, "type": 2, "phone": 3,
    "predstavlennia": 6, "kuzov": 7, "rik": 8, "probig": 9,
    "diagnostika": 10, "mynuli_roboty": 11, "zapys": 12, "proschannia": 13,
    "robota": 14, "instrukcii": 15, "rekomendacii": 16, "result": 17, 
    "ocinka": 18, "zapchastyny": 19, "comment": 20, "score": 21,
}

def write(rows):
    dest = Path(config.OUT_XLSX)
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(config.TEMPLATE_XLSX, dest)
    
    wb = openpyxl.load_workbook(dest)
    ws = wb['Лист1'] if 'Лист1' in wb.sheetnames else wb.active
    ws.cell(row=2, column=COLS["score"], value="Бали")

    r_idx = 3
    for r in rows:
        def put(key, val):
            return ws.cell(row=r_idx, column=COLS[key], value=val)

        put("date", r["date"])
        put("type", "Вхідний")
        
        p_val = r["phone"]
        try:
            p_val = int(p_val)
        except ValueError:
            pass
        put("phone", p_val)
        
        info = r["analysis"]
        for k in config.SCORE_KEYS:
            put(k, info[k])
            
        put("zapys", "так" if info["zapys"] else "")
        put("robota", info["robota"])
        put("rekomendacii", info["rekomendacii"])
        put("result", info["result"])
        
        # оцінка розмови
        if info.get("failed_analysis"):
            ocinka = "—"
            is_red = False
        else:
            # оцінка базується на балах з конфігу
            score = info.get("score")
            if score is not None and score != "" and int(score) >= config.GOOD_SCORE_THRESHOLD:
                ocinka = "добре"
                is_red = False
            else:
                ocinka = "погано"
                is_red = True
                
        put("ocinka", ocinka)
        put("zapchastyny", info["zapchastyny"])
        
        cmt_cell = put("comment", info["comment"])
        if is_red:
            cmt_cell.font = RED
            ws.cell(row=r_idx, column=COLS["ocinka"]).font = RED
            
        put("score", info["score"])
        r_idx += 1

    wb.save(dest)
    return str(dest)